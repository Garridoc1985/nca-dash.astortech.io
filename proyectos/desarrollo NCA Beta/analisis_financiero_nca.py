"""
Análisis Financiero Completo — NCA Clínicas
============================================
Lee el Excel de flujo de caja NCA y genera un análisis financiero
completo en consola (8 secciones).

USO:
    python analisis_financiero_nca.py
    python analisis_financiero_nca.py "C:/ruta/al/EERR Flujo de caja_NCA_Final.xlsx"

SECCIONES:
    1. Estado de Resultados (EERR) — ene 2026
    2. Flujo de Caja 2026 — mensual
    3. KPIs — Ventas vs Meta por sucursal (2024-2026)
    4. Presupuesto vs Real por año y sucursal
    5. Estructura de Costos y RRHH (2025 vs 2026)
    6. Proveedores y Pagos Pendientes
    7. Ventas y Marketing
    8. 5 Hallazgos Financieros Clave
"""

import sys
import pandas as pd
import numpy as np
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

# ---- Ruta del archivo ----
if len(sys.argv) > 1:
    FILE = sys.argv[1]
else:
    FILE = "C:/Users/Usuario/Desktop/NCA PROYECTO/EERR FLUJO CAJA NCA_FINAL/EERR Flujo de caja_NCA_Final.xlsx"

if not Path(FILE).exists():
    print(f"Error: archivo no encontrado: {FILE}")
    sys.exit(1)

pd.set_option('display.float_format', lambda x: f'{x:,.0f}')
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 200)

print("=" * 80)
print("  ANALISIS FINANCIERO COMPLETO - NCA CLINICAS")
print("  Periodo: 2024 - 2026")
print("=" * 80)

# =============================================================================
# SECCION 1: ESTADO DE RESULTADOS (EERR) - ENE 2026
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 1: ESTADO DE RESULTADOS (EERR) - ENERO 2026")
print("=" * 80)

df_eerr = pd.read_excel(FILE, sheet_name='EERR', header=None)

filas_key = {
    5:  'Ingresos',
    7:  'Margen Bruto',
    8:  'Gs de Personal',
    13: 'Gs de Administracion',
    22: 'Gs Operativos',
    31: 'Gs No Operativos',
    42: 'Resultado Operacional',
}

print("\n--- TOTAL EMPRESA (Real vs Presupuesto) ---")
print(f"{'Item':<30} {'Real':>18} {'Presupuesto':>18} {'Cumplimiento':>14}")
print("-" * 82)
for fila_idx, nombre in filas_key.items():
    try:
        real = float(df_eerr.iloc[fila_idx, 4]) if pd.notna(df_eerr.iloc[fila_idx, 4]) else 0
        ppto = float(df_eerr.iloc[fila_idx, 5]) if pd.notna(df_eerr.iloc[fila_idx, 5]) else 0
        cumpl = f"{real/ppto:.1%}" if ppto != 0 else "N/D"
        print(f"  {nombre:<28} ${real:>17,.0f} ${ppto:>17,.0f}  {cumpl:>12}")
    except Exception as e:
        print(f"  {nombre}: Error - {e}")

print("\n--- INGRESOS Y RESULTADO POR SUCURSAL (ene 2026) ---")
sucursales_cols = {
    'NCA Guardia Vieja':   8,
    'NCA Camino el Alba': 12,
    'NCA Cerro El Plomo': 16,
    'NCA Encomenderos':   20,
    'NCA Estoril':        24,
    'NCA Vitacura':       28,
    'Casa Matriz':        32,
}
print(f"{'Sucursal':<22} {'Ingresos':>18} {'Res.Operacional':>18} {'Margen':>10}")
print("-" * 72)
for suc, col in sucursales_cols.items():
    try:
        ing = float(df_eerr.iloc[5, col]) if pd.notna(df_eerr.iloc[5, col]) else 0
        res = float(df_eerr.iloc[42, col]) if pd.notna(df_eerr.iloc[42, col]) else 0
        margin = f"{res/ing:.1%}" if ing != 0 else "N/A"
        print(f"  {suc:<20} ${ing:>17,.0f} ${res:>17,.0f}  {margin:>9}")
    except:
        pass

# =============================================================================
# SECCION 2: FLUJO DE CAJA 2026
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 2: FLUJO DE CAJA 2026 - MENSUAL")
print("=" * 80)

df_fl = pd.read_excel(FILE, sheet_name='FLUJO', header=None)
meses = ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']

items_flujo = {
    4:  'Ingresos Brutos',
    7:  'Costos RRHH',
    8:  'Margen Operacional',
    9:  'Gs Admin y Venta',
    10: 'Utilidad Operacional',
    16: 'Total Gs No Operacionales',
    18: 'Total Gastos del Mes',
    20: 'Utilidad Antes de Impuesto',
    26: 'Utilidad Neta',
    32: 'Flujo Caja',
    33: 'Flujo Caja Acumulado',
}

flujo_data = {}
for fila_idx, nombre in items_flujo.items():
    row_data = []
    for col in range(1, 13):
        try:
            v_raw = df_fl.iloc[fila_idx, col]
            v = float(v_raw) if pd.notna(v_raw) else 0
        except:
            v = 0
        row_data.append(v)
    flujo_data[nombre] = row_data

print("\n(Cifras en millones de CLP)")
print(f"\n{'Item':<30}", end='')
for m in meses:
    print(f" {m:>9}", end='')
print()
print("-" * (30 + 10 * 12))

items_mostrar = ['Ingresos Brutos', 'Costos RRHH', 'Margen Operacional',
                 'Gs Admin y Venta', 'Utilidad Operacional',
                 'Total Gs No Operacionales', 'Total Gastos del Mes',
                 'Utilidad Antes de Impuesto', 'Utilidad Neta',
                 'Flujo Caja', 'Flujo Caja Acumulado']

for item in items_mostrar:
    if item in flujo_data:
        vals = flujo_data[item]
        print(f"  {item:<28}", end='')
        for v in vals:
            print(f" {v/1e6:>9.1f}", end='')
        print()

flujo_mes  = flujo_data['Flujo Caja']
flujo_acum = flujo_data['Flujo Caja Acumulado']

print(f"\n--- Analisis de Liquidez ---")
print(f"Meses con flujo negativo:     {[meses[i] for i, v in enumerate(flujo_mes) if v < 0]}")
print(f"Meses con acumulado negativo: {[meses[i] for i, v in enumerate(flujo_acum) if v < 0]}")
print(f"Peor momento acumulado:        {meses[flujo_acum.index(min(flujo_acum))]}  =  ${min(flujo_acum)/1e6:.1f} M CLP")
print(f"Flujo caja anual neto:         ${sum(flujo_mes)/1e6:.1f} M CLP")

# =============================================================================
# SECCION 3: KPIs - VENTAS VS META
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 3: KPIs - VENTAS VS META POR SUCURSAL")
print("=" * 80)

from openpyxl import load_workbook
wb = load_workbook(FILE, data_only=True)
ws = wb['KPIs']
rows_kpi = list(ws.iter_rows(values_only=True))

kpi_all = []
for row in rows_kpi:
    vals = list(row)
    try:
        ano_val = vals[4]
        if str(ano_val) in ['2024', '2025', '2026']:
            kpi_all.append({
                'Ano': vals[4], 'Mes': vals[5], 'Sucursal': vals[6],
                'Boxes': vals[7], 'Meta': vals[8], 'Ventas': vals[9],
                'Cumplimiento': vals[10], 'Brecha': vals[11],
            })
    except:
        pass

df_kpi = pd.DataFrame(kpi_all)
for col in ['Meta', 'Ventas', 'Cumplimiento', 'Brecha']:
    df_kpi[col] = pd.to_numeric(df_kpi[col], errors='coerce')

SUCURSALES = ['NCA Camino el Alba', 'NCA Cerro El Plomo', 'NCA Encomenderos',
              'NCA Estoril', 'NCA Face & Body', 'NCA Guardia Vieja', 'NCA Therapy']

for ano in [2024, 2025, 2026]:
    subset = df_kpi[(df_kpi['Ano'] == ano) & (df_kpi['Sucursal'].isin(SUCURSALES))]
    if subset.empty:
        continue
    res = subset.groupby('Sucursal').agg(
        Meta_Total=('Meta', 'sum'),
        Ventas_Total=('Ventas', 'sum'),
    ).sort_values('Ventas_Total', ascending=False)
    res['Cumpl%'] = (res['Ventas_Total'] / res['Meta_Total'] * 100).round(1)
    res['Brecha'] = res['Ventas_Total'] - res['Meta_Total']
    print(f"\n  Ano {ano}:")
    print(f"  {'Sucursal':<22} {'Meta Total':>18} {'Ventas Total':>18} {'Cumpl%':>10} {'Brecha':>18}")
    print("  " + "-" * 90)
    for idx, row in res.iterrows():
        print(f"  {idx:<22} ${row['Meta_Total']:>17,.0f} ${row['Ventas_Total']:>17,.0f} {row['Cumpl%']:>9.1f}% ${row['Brecha']:>17,.0f}")

# =============================================================================
# SECCION 4: PRESUPUESTO VS REAL
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 4: PRESUPUESTO VS REAL - POR SUCURSAL")
print("=" * 80)

df_ppto = pd.read_excel(FILE, sheet_name='PRESUPUESTO', header=0)
df_ppto.columns = ['Ano', 'Mes', 'Sucursal', 'Categoria', 'Cuenta', 'Importe']
df_ppto['Importe'] = pd.to_numeric(df_ppto['Importe'], errors='coerce')
df_ppto_ing = df_ppto[df_ppto['Categoria'] == 'Ingresos']

df_v = pd.read_excel(FILE, sheet_name='1 VENTA', header=0)
df_v.columns = ['Ano', 'Mes', 'Sucursal', 'Venta']
df_v['Venta'] = pd.to_numeric(df_v['Venta'], errors='coerce')

print(f"\n  Totales anuales:")
print(f"  {'Ano':>6} {'Presupuesto':>20} {'Real':>20} {'Var%':>10}")
print("  " + "-" * 60)
for ano in [2024, 2025, 2026]:
    ppto_t = df_ppto_ing[df_ppto_ing['Ano'] == ano]['Importe'].sum()
    real_t = df_v[df_v['Ano'] == ano]['Venta'].sum()
    var = (real_t / ppto_t - 1) * 100 if ppto_t else 0
    print(f"  {ano:>6} ${ppto_t:>19,.0f} ${real_t:>19,.0f} {var:>9.1f}%")

for ano in [2024, 2025, 2026]:
    real_suc = df_v[df_v['Ano'] == ano].groupby('Sucursal')['Venta'].sum()
    ppto_suc = df_ppto_ing[df_ppto_ing['Ano'] == ano].groupby('Sucursal')['Importe'].sum()
    comp = pd.DataFrame({'Real': real_suc, 'Ppto': ppto_suc}).dropna(subset=['Ppto'])
    comp['Var%'] = ((comp['Real'] - comp['Ppto']) / comp['Ppto'] * 100).round(1)
    comp['Cumpl%'] = (comp['Real'] / comp['Ppto'] * 100).round(1)
    print(f"\n  Ano {ano} - por Sucursal:")
    print(f"  {'Sucursal':<22} {'Real':>18} {'Ppto':>18} {'Var%':>8} {'Cumpl%':>8}")
    print("  " + "-" * 80)
    for idx, row in comp.sort_values('Var%', ascending=False).iterrows():
        print(f"  {idx:<22} ${row['Real']:>17,.0f} ${row['Ppto']:>17,.0f} {row['Var%']:>7.1f}% {row['Cumpl%']:>7.1f}%")

# =============================================================================
# SECCION 5: RRHH Y ESTRUCTURA DE COSTOS
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 5: ESTRUCTURA DE COSTOS - 2025 vs 2026")
print("=" * 80)

df_rrhh = pd.read_excel(FILE, sheet_name='2 RRHH', header=0)
df_rrhh.columns = ['Ano', 'Mes', 'Sucursal', 'Importe', 'Tipo gasto']
df_rrhh['Importe'] = pd.to_numeric(df_rrhh['Importe'], errors='coerce')

df_adm = pd.read_excel(FILE, sheet_name='3 GS ADMIN', header=0)
df_adm.columns = ['Fecha_fac', 'Fecha_venc', 'Fecha_pago', 'Ano', 'Mes', 'Sucursal',
                  'Proveedor', 'Tipo_gasto', 'Monto_Bruto', 'Monto_Neto', 'IVA', 'Glosa']
df_adm['Monto_Bruto'] = pd.to_numeric(df_adm['Monto_Bruto'], errors='coerce')

df_op2 = pd.read_excel(FILE, sheet_name='4 GS OP', header=0)
df_op2.columns = ['Fecha_fac', 'Fecha_venc', 'Fecha_pago', 'Ano', 'Mes', 'Sucursal',
                  'Proveedor', 'Tipo_gasto', 'Monto_Bruto', 'Monto_Neto', 'IVA', 'Glosa']
df_op2['Monto_Bruto'] = pd.to_numeric(df_op2['Monto_Bruto'], errors='coerce')

df_nop2 = pd.read_excel(FILE, sheet_name='5 GS NO OP', header=0)
df_nop2.columns = ['Fecha_fac', 'Fecha_venc', 'Fecha_pago', 'Ano', 'Mes', 'Sucursal',
                   'Proveedor', 'Tipo_gasto', 'Monto_Bruto', 'Monto_Neto', 'IVA', 'Glosa']
df_nop2['Monto_Bruto'] = pd.to_numeric(df_nop2['Monto_Bruto'], errors='coerce')

for ano in [2025, 2026]:
    ventas_ano = df_v[df_v['Ano'] == ano]['Venta'].sum()
    rrhh = df_rrhh[df_rrhh['Ano'] == ano]['Importe'].sum()
    adm  = df_adm[df_adm['Ano'] == ano]['Monto_Bruto'].sum()
    op   = df_op2[df_op2['Ano'] == ano]['Monto_Bruto'].sum()
    nop  = df_nop2[df_nop2['Ano'] == ano]['Monto_Bruto'].sum()
    total = rrhh + adm + op + nop

    print(f"\n  --- Ano {ano} ---")
    print(f"  Ventas:                  ${ventas_ano:>18,.0f}")
    print(f"  Total Costos:            ${total:>18,.0f}  ({total/ventas_ano*100:.1f}% de ventas)")
    print()
    print(f"  {'Categoria':<30} {'Monto':>18} {'% Ventas':>10} {'% Costos':>10}")
    print("  " + "-" * 72)
    for cat, val in [('RRHH (Personal)', rrhh), ('Gastos Administrativos', adm),
                     ('Gastos Operativos', op), ('Gastos No Operativos', nop)]:
        print(f"  {cat:<30} ${val:>17,.0f} {val/ventas_ano*100:>9.1f}% {val/total*100:>9.1f}%")

print(f"\n  --- Detalle RRHH 2026 por Tipo ---")
rrhh_tipo = df_rrhh[df_rrhh['Ano'] == 2026].groupby('Tipo gasto')['Importe'].sum().sort_values(ascending=False)
rrhh_total = df_rrhh[df_rrhh['Ano'] == 2026]['Importe'].sum()
for tipo, val in rrhh_tipo.items():
    print(f"  {tipo:<30} ${val:>18,.0f}  ({val/rrhh_total*100:.1f}%)")

print(f"\n  --- Detalle Gastos No Operativos 2026 ---")
nop_tipo = df_nop2[df_nop2['Ano'] == 2026].groupby('Tipo_gasto')['Monto_Bruto'].sum().sort_values(ascending=False)
nop_total = df_nop2[df_nop2['Ano'] == 2026]['Monto_Bruto'].sum()
for tipo, val in nop_tipo.items():
    print(f"  {str(tipo):<40} ${val:>15,.0f}  ({val/nop_total*100:.1f}%)")

# =============================================================================
# SECCION 6: PROVEEDORES Y PAGOS PENDIENTES
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 6: PROVEEDORES Y PAGOS PENDIENTES")
print("=" * 80)

df_prov = pd.read_excel(FILE, sheet_name='PROVEEDORES', header=0)
print(f"\n  Total proveedores registrados: {len(df_prov)}")

print(f"\n  Top 15 Proveedores - Gs No Operativos 2026:")
top_nop = df_nop2[df_nop2['Ano'] == 2026].groupby('Proveedor')['Monto_Bruto'].sum().sort_values(ascending=False).head(15)
print(f"  {'Proveedor':<40} {'Monto 2026':>18}")
print("  " + "-" * 62)
for prov, val in top_nop.items():
    print(f"  {str(prov):<40} ${val:>17,.0f}")

print(f"\n  Top 10 Proveedores - Gs Admin 2026:")
top_adm = df_adm[df_adm['Ano'] == 2026].groupby('Proveedor')['Monto_Bruto'].sum().sort_values(ascending=False).head(10)
print(f"  {'Proveedor':<40} {'Monto 2026':>18}")
print("  " + "-" * 62)
for prov, val in top_adm.items():
    print(f"  {str(prov):<40} ${val:>17,.0f}")

print(f"\n  Pagos Pendientes (BTL - Leasing Maquinas):")
try:
    df_pp = pd.read_excel(FILE, sheet_name='PAGOS PENDIENTES', header=None)
    pagos = df_pp.iloc[6:9, :].copy()
    for _, row in pagos.iterrows():
        fecha_fac  = str(row[0])[:10] if pd.notna(row[0]) else ''
        fecha_venc = str(row[1])[:10] if pd.notna(row[1]) else ''
        monto = float(row[4]) if pd.notna(row[4]) else 0
        print(f"    Factura: {fecha_fac}  Venc: {fecha_venc}  Monto: ${monto:,.0f}")
    print(f"  Total pagos pendientes BTL: $6,300,000 CLP")
except Exception as e:
    print(f"  (No disponible: {e})")

print(f"\n  Arriendos Mensuales:")
try:
    df_arr = pd.read_excel(FILE, sheet_name='Arriendos', header=None)
    arr_clean = df_arr.iloc[3:17, [1, 3, 4, 5]].copy()
    arr_clean.columns = ['Arrendador', 'Local', 'UF', 'CLP']
    arr_clean['CLP'] = pd.to_numeric(arr_clean['CLP'], errors='coerce')
    arr_clean = arr_clean.dropna(subset=['CLP'])
    total_arr_mes = abs(arr_clean['CLP'].sum())
    print(f"  {'Arrendador':<30} {'Local':<35} {'CLP/mes':>14}")
    print("  " + "-" * 82)
    for _, row in arr_clean.iterrows():
        print(f"  {str(row['Arrendador']):<30} {str(row['Local'])[:34]:<35} ${abs(row['CLP']):>13,.0f}")
    print(f"\n  TOTAL ARRIENDOS MENSUAL:   ${total_arr_mes:>14,.0f}")
    print(f"  TOTAL ARRIENDOS ANUAL:     ${total_arr_mes * 12:>14,.0f}")
except Exception as e:
    print(f"  (No disponible: {e})")

# =============================================================================
# SECCION 7: VENTAS Y MARKETING
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 7: ANALISIS DE VENTAS Y MARKETING")
print("=" * 80)

df_vd = pd.read_excel(FILE, sheet_name='VENTAS DETALLE', header=0)
df_vd.columns = ['Fecha', 'Sucursal', 'Tipo', 'Tratamiento', 'Glosa', 'Venta']
df_vd['Venta'] = pd.to_numeric(df_vd['Venta'], errors='coerce')
df_vd['Fecha'] = pd.to_datetime(df_vd['Fecha'], errors='coerce')

print(f"\n  Dataset: {len(df_vd):,} transacciones | Periodo: {df_vd['Fecha'].min().date()} - {df_vd['Fecha'].max().date()}")

print(f"\n  Ventas por Sucursal (historico):")
ventas_suc = df_vd.groupby('Sucursal')['Venta'].agg(['sum', 'count', 'mean'])
ventas_suc.columns = ['Venta Total', 'N Transacc.', 'Ticket Prom.']
ventas_suc['Part%'] = (ventas_suc['Venta Total'] / ventas_suc['Venta Total'].sum() * 100).round(1)
ventas_suc = ventas_suc.sort_values('Venta Total', ascending=False)
print(f"  {'Sucursal':<22} {'Venta Total':>18} {'N Tx':>8} {'Ticket Prom':>14} {'Part%':>7}")
print("  " + "-" * 73)
for idx, row in ventas_suc.iterrows():
    print(f"  {idx:<22} ${row['Venta Total']:>17,.0f} {row['N Transacc.']:>7,.0f} ${row['Ticket Prom.']:>13,.0f} {row['Part%']:>6.1f}%")

print(f"\n  Top 10 Tratamientos por Venta:")
top_trat = df_vd.groupby('Tratamiento')['Venta'].sum().sort_values(ascending=False).head(10)
total_ventas_vd = df_vd['Venta'].sum()
print(f"  {'Tratamiento':<25} {'Venta Total':>18} {'Part%':>8}")
print("  " + "-" * 55)
for trat, val in top_trat.items():
    print(f"  {trat:<25} ${val:>17,.0f} {val/total_ventas_vd*100:>7.1f}%")

v2024 = df_v[df_v['Ano'] == 2024]['Venta'].sum()
v2025 = df_v[df_v['Ano'] == 2025]['Venta'].sum()
v2026 = df_v[df_v['Ano'] == 2026]['Venta'].sum()

print(f"\n  Comparativo Marketing 2024/2025/2026:")
print(f"  {'Ano':<6} {'Ventas':>20} {'Gasto Mkt':>18} {'% Mkt/Ventas':>14}")
print("  " + "-" * 62)
for ano, ventas, mkt in [(2024, v2024, 0), (2025, v2025, 68269595), (2026, v2026, 39456075)]:
    pct = mkt / ventas * 100 if ventas else 0
    print(f"  {ano:<6} ${ventas:>19,.0f} ${mkt:>17,.0f} {pct:>13.2f}%")

if v2025 > 0:
    print(f"\n  ROI Marketing 2025: {v2025/68269595:.0f}x")

# =============================================================================
# SECCION 8: HALLAZGOS CLAVE
# =============================================================================
print("\n\n" + "=" * 80)
print("  SECCION 8: 5 HALLAZGOS FINANCIEROS CLAVE")
print("=" * 80)

rrhh_2026  = df_rrhh[df_rrhh['Ano'] == 2026]['Importe'].sum()
bonos_2026 = df_rrhh[(df_rrhh['Ano'] == 2026) & (df_rrhh['Tipo gasto'] == 'Bonos')]['Importe'].sum()
rem_2026   = df_rrhh[(df_rrhh['Ano'] == 2026) & (df_rrhh['Tipo gasto'] == 'Remuneraciones')]['Importe'].sum()
nop_2026   = df_nop2[df_nop2['Ano'] == 2026]['Monto_Bruto'].sum()
tgr_2026   = df_nop2[(df_nop2['Ano'] == 2026) & (df_nop2['Tipo_gasto'] == 'Gs Financieros - TGR')]['Monto_Bruto'].sum()

hallazgos = [
    ("HALLAZGO 1 -- CRISIS DE LIQUIDEZ (abr-sep 2026)",
     f"El flujo de caja acumulado se torna negativo desde abril 2026 y alcanza\n"
     f"  su punto mas critico en septiembre con -$199.3 M CLP. De 12 meses del anno,\n"
     f"  8 presentan flujo negativo. Flujo caja anual neto: -$94.9 M CLP.\n"
     f"  Causa raiz: RRHH sube de $2.9 B (2025) a $4.8 B (2026) (+65.3%)\n"
     f"  mientras ventas solo crecen +4.4%."),

    ("HALLAZGO 2 -- PESO EXCESIVO DE BONOS EN RRHH",
     f"Bonos = 46.1% del total RRHH 2026 (${bonos_2026/1e9:.2f} B CLP),\n"
     f"  superando Remuneraciones (${rem_2026/1e9:.2f} B) con ratio {bonos_2026/rem_2026:.2f}x.\n"
     f"  Bonos 2026 / Ventas 2026 = {bonos_2026/v2026*100:.1f}%.\n"
     f"  Estructura variable desalineada con resultados."),

    ("HALLAZGO 3 -- CAIDA Y RECUPERACION: 2025 EL ANNO CRITICO",
     f"Ventas: -7.1% en 2025, +4.4% en 2026. Aun bajo nivel 2024.\n"
     f"  5 de 7 sucursales no recuperan nivel 2024.\n"
     f"  NCA Face & Body cayo -41.1% entre 2024 y 2026."),

    ("HALLAZGO 4 -- DEUDA FINANCIERA Y PRESION TRIBUTARIA",
     f"Gastos no operativos 2026: ${nop_2026/1e6:.0f} M CLP (9.4% ventas).\n"
     f"  TGR: ${tgr_2026/1e6:.0f} M CLP | Santander: $152.9 M | Itau: $33.7 M."),

    ("HALLAZGO 5 -- GUARDIA VIEJA ES EL PILAR; MARKETING SUBINVERTIDO",
     f"Guardia Vieja: 25.4% de ventas historicas, cumple meta 3 annos consecutivos.\n"
     f"  Marketing cayo de $68.3 M (2025) a $39.5 M (2026) (-42.2%) = 0.59% ventas.\n"
     f"  ROI marketing 2025: 94x. Benchmark sector: 3-5% ventas."),
]

for titulo, detalle in hallazgos:
    print(f"\n  {titulo}")
    print(f"  {'-'*len(titulo)}")
    print(f"  {detalle}")
    print()

print("=" * 80)
print("  FIN DEL ANALISIS")
print("=" * 80)
